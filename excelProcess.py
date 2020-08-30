from openpyxl import load_workbook
import os
import re
from unittest import TestCase, main

sourceFolder = "D:\wanghuayang\python"
excelName = "HX11_M2 boxcar_CM01_20200721.xlsx"
targetFile = os.path.join(sourceFolder, excelName)
targetSheet = "CM01-ECUlist"

colIndex = (
    2,  #"B" for number
    3,  #"C" for name
    4,  #"D" for variant
    6,  #"F" for owner
    11, #"K" for assemly number
    12, #"L" for HW version
)
colName = (
    "number",
    "name",
    "variant",
    "owner",
    "assemblyNr",
    "HWVer",
)

isNumber = re.compile(r'\d+')

# def getEcuList(min_row, max_row, )

class ECU:
    def __init__(self, *args, **kwargs):
        self.name=kwargs["name"]
        self.variant = kwargs["variant"]
        self.owner = kwargs["owner"]
        self.description = kwargs["owner"]
        self.assembelyNumber = kwargs["assemblyNr"]
        self.supplier = None
    
    def setVariant(self, variant):
        self.variant = variant
    
    def setOwner(self, owner):
        self.owner=owner

class Configuration:
    def __init__(self, *args, **kwargs):
        self.name = kwargs['name']
        self.excelSetup(kwargs['targetFile'], kwargs['targetSheet'])
        self.ecuList = []

    def excelSetup(self, targetFile, targetSheet):
        self.ecuBook = load_workbook(targetFile)
        self.ecuSheet = self.ecuBook.get_sheet_by_name(targetSheet)

    def getEcuListFromExcel(self, minRow, maxRow, minCol, maxCol):
        ecuList = []
        # ecu = {}
        print("Excel area of Min Row: {0}, Max Row: {1}, Min Col: {0}, Max Col: {1}".format(minRow, maxRow, minCol, maxCol))
        print(self.ecuSheet)
        for row in self.ecuSheet.iter_rows(min_row=minRow, max_row=maxRow, min_col=minCol, max_col=maxCol) :
            ecuInfo = []
            # print(row)
            for index in colIndex:
                # print("index: {0}, value: {1}".format(index-1, row[index-1].value))
                ecuInfo.append(row[index-1].value)
                # print(row[index-1].value)
            # print(ecuInfo)
            ecuInfo = tuple(ecuInfo)
            ecu = dict(zip(colName, ecuInfo)) # combine list into dictionary
            if ecu["name"] and isNumber.match(ecu["number"]):
                ecuList.append(ecu)
        self.ecuList = ecuList
        self.storeEcu()
        return ecuList

    def storeEcu(self):
        Storage = {}
        # Storage["name"] = {}
        if self.ecuList:
            for ecuInfo in self.ecuList:
                # print(ecuInfo)
                # print(ecuInfo["name"])
                Storage.setdefault(ecuInfo["name"], []).append(ecuInfo)
            self.ecuStorage = Storage
        else:
            print("Please manually execute method 'getEcuListFromExcel' in advance!")
        return Storage

    def lookupEcu(self, name):
        return self.ecuStorage.get(name)



class TestExcelProcess(TestCase):
    def setUp(self):
        cm01 = Configuration(name="CM01", targetFile=targetFile, targetSheet=targetSheet)
        excelArea = [   4,  #min Row 
                        94,  #max Row
                        0,  #min Column
                        13  #max Column
                        ]
        cm01.getEcuListFromExcel(*excelArea)
        self.conf = cm01

    def testExcelRead(self):
        self.assertIsNotNone(self.conf.ecuList)
        # print(self.ecuList[0])

    def testStorage(self):
        self.assertGreaterEqual(len(self.conf.ecuStorage["CEM"]),1)

    def testLookupEcu(self):
        ecu = self.conf.lookupEcu("CEM")
        print("Check the ecu content: {0}".format(ecu))
        self.assertIn("CEM", ecu[0].values())

    def tearDown(self):
        pass


if __name__ == "__main__":
    main()
